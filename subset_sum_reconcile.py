# -*- coding: utf-8 -*-
"""
Accounting reconciliation tool — Subset Sum with branch-and-bound pruning.

Reads an Excel file (sheet 'Input') containing a list of transactions and a
target sum, and finds up to 10 subsets of transactions whose sum equals the
target exactly. The shortest solutions are returned first. The output is
written to the 'Output' sheet in the same file.

All computation is done on integer cents (value * 100) to avoid
floating-point rounding errors — critical in an accounting context.

The strategy: split the problem into a cents problem and a whole-unit problem.
  • "Round" transactions (0 cents) do not contribute to the cents part of the sum.
  • Therefore the cents part must be covered by *only* the non-round transactions.
  • Step 1: find subsets of the non-round transactions whose cents sum ≡ the
           target cents (mod 100). These are the "candidates".
  • Step 2: for each candidate, complete the remainder (always a whole-unit
           amount) using a standard subset-sum search over the round transactions only.

Usage:
    python subset_sum_reconcile.py <path-to-file.xlsx>
"""

import sys
import os
import time
import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


MAX_SOLUTIONS = 10  # stop after 10 solutions — beyond that you need extra identifying info

# --- Excel workbook styling (headers, borders, alignment) --- #
_HEADER_FILL = PatternFill("solid", fgColor="2F6BE6")       # accent blue
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=16, color="1F2430")
_LABEL_FONT = Font(bold=True, color="1F2430")
_MUTED_FONT = Font(color="6B7280")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_MONEY = '#,##0.00'

# Status text for the 'Output' sheet.
STATUS_TEXT = {
    "complete":  "Search complete — these are all solutions up to the item limit.",
    "cap":       "Stopped after {n} solutions (there may be more — narrow the search or add identifying info).",
    "time":      "Stopped on the time limit — these are the solutions found so far.",
    "interrupt": "Interrupted manually — these are the solutions found so far.",
}

# Status text for the console.
STATUS_TEXT_EN = {
    "complete":  "Search complete - these are all solutions up to the item limit.",
    "cap":       "Stopped after {n} solutions (there may be more - narrow the search or add identifying info).",
    "time":      "Stopped on time limit - these are the solutions found so far.",
    "interrupt": "Interrupted manually - these are the solutions found so far.",
}


class _StopSearch(Exception):
    """Internal signal to stop the whole search immediately (time limit / Ctrl+C)."""


def to_cents(value):
    """Convert a value (float) to an integer number of cents (minor units)."""
    return int(round(float(value) * 100))


# --------------------------------------------------------------------------- #
#  Search control: progress, time limit, and solution collection with dedup.
# --------------------------------------------------------------------------- #
class SearchState:
    """Holds the search state: time, branch counter, and the collected solutions."""

    def __init__(self, max_seconds, cancel_event=None, progress_callback=None,
                 max_solutions=MAX_SOLUTIONS):
        self.start = time.time()
        self.deadline = (self.start + max_seconds) if max_seconds else None
        self.last_report = self.start
        self.max_solutions = max_solutions     # solution quota (configurable)
        self.nodes = 0
        self.solutions = []   # list of tuples of cent values, sorted high to low
        self.seen = set()     # canonical form of solutions already collected (dedup)
        self.stop_reason = None
        # Optional hooks for the GUI:
        self.cancel_event = cancel_event          # threading.Event for manual stop
        self.progress_callback = progress_callback  # called instead of print while running

    def tick(self):
        """Called at every DFS node: counts branches, checks stop/time limit, reports progress."""
        self.nodes += 1
        if self.nodes & 0x3FFF == 0:          # every 16384 nodes
            now = time.time()
            if self.cancel_event is not None and self.cancel_event.is_set():
                self.stop_reason = "interrupt"
                raise _StopSearch
            if self.deadline and now > self.deadline:
                self.stop_reason = "time"
                raise _StopSearch
            interval = 0.4 if self.progress_callback else 2.0
            if now - self.last_report >= interval:
                self.last_report = now
                if self.progress_callback is not None:
                    self.progress_callback(self.nodes, len(self.solutions),
                                           now - self.start)
                else:
                    print(f"  ... working: {self.nodes:,} branches, "
                          f"{len(self.solutions)} solutions, {now - self.start:.0f}s",
                          flush=True)

    def record(self, values, target_total):
        """Adds a solution if its value list has not been seen yet. Returns True if added."""
        canon = tuple(sorted(values, reverse=True))
        assert sum(canon) == target_total, "Solution sum does not match the target — bug!"
        if canon in self.seen:
            return False
        self.seen.add(canon)
        self.solutions.append(canon)
        return True


# --------------------------------------------------------------------------- #
#  Step 0: read the input.
# --------------------------------------------------------------------------- #
def read_input(filepath):
    """
    Reads the 'Input' sheet.
    Returns (transactions, target, k_max) — monetary values (float), k_max as an integer.
    """
    if not os.path.isfile(filepath):
        raise SystemExit(f"ERROR: file not found: {filepath}")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except PermissionError:
        raise SystemExit("ERROR: cannot open the file (maybe open in Excel). Close it and try again.")
    except Exception as e:
        raise SystemExit(f"ERROR reading the file: {e}")

    if "Input" not in wb.sheetnames:
        raise SystemExit(f"ERROR: input sheet 'Input' not found. Sheets present: {wb.sheetnames}")
    ws = wb["Input"]

    transactions = []
    skipped = 0
    for row in range(2, ws.max_row + 1):          # from A2 down
        v = ws.cell(row, 1).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        if isinstance(v, bool):                    # bool is a subclass of int — skip it
            skipped += 1
            continue
        if isinstance(v, (int, float)):
            transactions.append(float(v))
        else:
            skipped += 1
            print(f"  Warning: non-numeric value in cell A{row} ('{v}') - skipped.")

    if not transactions:
        raise SystemExit("ERROR: no numeric transactions found in column A (from A2 down).")
    if skipped:
        print(f"  Total skipped (non-numeric): {skipped}.")

    target = ws["D2"].value
    if target is None or (isinstance(target, str) and target.strip() == ""):
        raise SystemExit("ERROR: target sum (cell D2) is empty.")
    if isinstance(target, bool) or not isinstance(target, (int, float)):
        raise SystemExit(f"ERROR: target sum (D2) is not a number: '{target}'.")
    # target == 0 is allowed: finds subsets that cancel out (e.g. +500 and -500).

    kmax_raw = ws["D3"].value
    if kmax_raw is None or (isinstance(kmax_raw, str) and str(kmax_raw).strip() == ""):
        k_max = 8
    else:
        try:
            k_max = int(kmax_raw)
            if k_max < 1:
                k_max = 8
        except (ValueError, TypeError):
            print(f"  Warning: item limit (D3) invalid ('{kmax_raw}') - using 8.")
            k_max = 8

    return transactions, float(target), k_max


# --------------------------------------------------------------------------- #
#  Step 1: pre-filtering and splitting into round / non-round.
# --------------------------------------------------------------------------- #
def preprocess(tx_cents, target_total):
    """
    Filters out irrelevant transactions and splits into 'round' (0 cents) and 'non-round'.
    Input in cents (int). Returns (rounded, non_rounded, stats), each group sorted
    high to low (to enable early upper-bound pruning in the DFS).
    """
    rounded, non_rounded = [], []
    removed_large = removed_nonpos = 0
    for c in tx_cents:
        if c <= 0:
            removed_nonpos += 1
            continue
        if c > target_total:                       # larger than the target — can't be in a solution
            removed_large += 1
            continue
        (rounded if c % 100 == 0 else non_rounded).append(c)

    rounded.sort(reverse=True)
    non_rounded.sort(reverse=True)

    stats = {
        "total_in": len(tx_cents),
        "removed_large": removed_large,
        "removed_nonpos": removed_nonpos,
        "kept": len(rounded) + len(non_rounded),
        "n_rounded": len(rounded),
        "n_non_rounded": len(non_rounded),
    }
    return rounded, non_rounded, stats


# --------------------------------------------------------------------------- #
#  Step 2: the cents problem — finding candidates among the non-round transactions.
# --------------------------------------------------------------------------- #
def find_modular_candidates(non_rounded, target_cents, exact_len, state, lo, hi):
    """
    Finds subsets of the non-round transactions *of exactly length exact_len* whose
    cents sum ≡ target_cents (mod 100), with a total sum in the range [lo, hi].
    **Generator**: yields one candidate at a time — O(K) memory, without piling up
    millions of candidates in RAM (avoids Out-Of-Memory crashes). The caller checks
    each candidate immediately.

    The range [lo, hi] is supplied externally and is tight: the caller knows it must
    complete exactly `needed` round transactions, so the candidate's sum must be between
    target - (sum of the `needed` largest round values) and target - (sum of the `needed` smallest).
    At each node we prune if the range is no longer reachable given the m remaining items
    to pick (the m largest / m smallest). This prunes well even when the round pool is large.
    """
    if exact_len == 0:                              # the empty set (sum 0)
        if target_cents == 0 and lo <= 0 <= hi:
            yield ((), 0)
        return

    n = len(non_rounded)
    prefix = [0] * (n + 1)                          # prefix[i] = sum of non_rounded[:i]
    for i in range(n):
        prefix[i + 1] = prefix[i] + non_rounded[i]

    chosen = []

    def dfs(start, runsum, count):
        state.tick()
        if count == exact_len:                      # reached the requested length
            if lo <= runsum <= hi and runsum % 100 == target_cents:
                yield (tuple(chosen), runsum)       # emit in real time (no piling up in a list)
            return
        m = exact_len - count                       # how many more items still to pick
        if n - start < m:                           # not enough items left
            return
        max_add = prefix[start + m] - prefix[start]      # sum of the m largest remaining
        if runsum + max_add < lo:                   # even the best can't reach lo
            return
        min_add = prefix[n] - prefix[n - m]              # sum of the m smallest pickable
        if runsum + min_add > hi:                   # even the smallest exceeds hi
            return
        for i in range(start, n):
            v = non_rounded[i]
            if runsum + v > hi:                     # single item too large — skip it
                continue
            if i > start and non_rounded[i] == non_rounded[i - 1]:
                continue                            # skip duplicate at the same level
            chosen.append(v)
            yield from dfs(i + 1, runsum + v, count + 1)
            chosen.pop()

    yield from dfs(0, 0, 0)


# --------------------------------------------------------------------------- #
#  Subset-sum search with pruning (completes the whole-unit amount from the round pool).
# --------------------------------------------------------------------------- #
def find_subset_sum(values, target, max_len, state, exact_len=None, limit=None,
                    prefix=None):
    """
    Finds subsets of `values` whose sum equals `target` exactly.
    `values` is sorted high to low (int cents). If exact_len is given — only sets
    of that length; otherwise up to max_len. `limit` stops after a number of results (to save work).
    `prefix` — a precomputed prefix-sums array of `values` (avoids recomputation across many calls).
    Each multiset of values appears only once. Returns a list of tuples.
    """
    n = len(values)
    if prefix is None:                              # prefix[i] = sum of values[:i]
        prefix = [0] * (n + 1)
        for i in range(n):
            prefix[i + 1] = prefix[i] + values[i]

    cap_len = exact_len if exact_len is not None else max_len
    results = []
    chosen = []

    class _Enough(Exception):
        pass

    def dfs(start, runsum, count):
        state.tick()
        if runsum == target and (exact_len is None or count == exact_len):
            results.append(tuple(chosen))
            if limit is not None and len(results) >= limit:
                raise _Enough
            return                                  # positive values — can't extend to the same sum
        if count >= cap_len:                        # depth pruning
            return
        if start >= n:                              # end-of-list pruning
            return
        if exact_len is not None:                   # we know exactly how many items are missing — tight bound
            m = exact_len - count
            if n - start < m:                       # not enough items left
                return
            if runsum + (prefix[start + m] - prefix[start]) < target:   # the m largest aren't enough
                return
            if runsum + (prefix[n] - prefix[n - m]) > target:           # the m smallest already exceed
                return
        else:                                       # free length — ordinary lower bound
            if runsum + (prefix[n] - prefix[start]) < target:
                return
        for i in range(start, n):
            v = values[i]
            if runsum + v > target:                 # upper-bound pruning
                continue
            if i > start and values[i] == values[i - 1]:
                continue                            # skip duplicate at the same level
            chosen.append(v)
            dfs(i + 1, runsum + v, count + 1)
            chosen.pop()

    try:
        dfs(0, 0, 0)
    except _Enough:
        pass
    return results


def format_amounts(cents_seq):
    """Display string for solution values: '600.50  +  400.00', and with negatives '… -  50.30'."""
    vals = sorted(cents_seq, reverse=True)
    s = f"{vals[0] / 100:,.2f}"
    for c in vals[1:]:
        if c < 0:
            s += f"  -  {abs(c) / 100:,.2f}"
        else:
            s += f"  +  {c / 100:,.2f}"
    return s


def _finalize(state, stats, target_total):
    """Determines status, sorts shortest solutions first, and fills in the stats. Returns (solutions, stats)."""
    if state.stop_reason in ("time", "interrupt"):
        status = state.stop_reason
    elif len(state.solutions) >= state.max_solutions:
        status = "cap"
    else:
        status = "complete"
    solutions = sorted(state.solutions,
                       key=lambda s: (len(s), tuple(-x for x in s)))
    stats.update(status=status,
                 status_text=STATUS_TEXT[status].format(n=state.max_solutions),
                 nodes=state.nodes, elapsed=time.time() - state.start,
                 target_total=target_total, max_solutions=state.max_solutions)
    return solutions, stats


# --------------------------------------------------------------------------- #
#  General path: mixed-sign values (including negatives) and any target.
# --------------------------------------------------------------------------- #
def _solve_general(tx_cents, target_total, k_max, state):
    """
    The positive path relies on monotonicity (all values positive) for pruning. When there
    are negatives this no longer holds, so here we search by increasing length with a tight
    two-sided bound: when picking m additional items from those remaining, the added sum is
    bounded between the sum of the m smallest and the sum of the m largest — which prunes
    branches that can't reach the target.
    """
    vals = sorted((c for c in tx_cents if c != 0), reverse=True)
    n = len(vals)
    prefix = [0] * (n + 1)
    for i in range(n):
        prefix[i + 1] = prefix[i] + vals[i]

    n_round = sum(1 for c in vals if c % 100 == 0)
    stats = {
        "total_in": len(tx_cents),
        "removed_large": 0,
        "removed_nonpos": sum(1 for c in tx_cents if c == 0),
        "kept": n,
        "n_rounded": n_round,
        "n_non_rounded": n - n_round,
        "n_candidates": 0,
    }

    chosen = []

    def dfs(start, runsum, count, K):
        state.tick()
        if len(state.solutions) >= state.max_solutions:
            return
        if count == K:
            if runsum == target_total:
                state.record(tuple(chosen), target_total)
            return
        m = K - count                              # m more items to pick
        if n - start < m:                          # not enough left
            return
        need = target_total - runsum
        max_add = prefix[start + m] - prefix[start]   # sum of the m largest remaining
        min_add = prefix[n] - prefix[n - m]           # sum of the m smallest overall
        if need > max_add or need < min_add:       # two-sided pruning
            return
        for i in range(start, n):
            if i > start and vals[i] == vals[i - 1]:
                continue                           # skip duplicate at the same level
            chosen.append(vals[i])
            dfs(i + 1, runsum + vals[i], count + 1, K)
            chosen.pop()
            if len(state.solutions) >= state.max_solutions:
                return

    try:
        for K in range(1, k_max + 1):              # shortest solutions first
            dfs(0, 0, 0, K)
            if len(state.solutions) >= state.max_solutions:
                break
    except (_StopSearch, KeyboardInterrupt):
        if state.stop_reason is None:
            state.stop_reason = "interrupt"

    return _finalize(state, stats, target_total)


# --------------------------------------------------------------------------- #
#  The core: combining the steps with iterative deepening (shortest solutions first).
# --------------------------------------------------------------------------- #
def solve(transactions, target, k_max, max_seconds,
          cancel_event=None, progress_callback=None, max_results=None):
    """
    Returns (solutions, stats). solutions is a list of tuples of cent values,
    sorted by item count ascending, up to max_results solutions (default 10).

    cancel_event / progress_callback — optional hooks for a GUI:
      cancel_event       — threading.Event; when set, the search stops and the
                           results found so far are returned.
      progress_callback  — function(nodes, n_solutions, elapsed) called
                           occasionally while running instead of printing to the console.
      max_results        — solution quota; if None -> MAX_SOLUTIONS (10).
    """
    target_total = to_cents(target)
    tx_cents = [to_cents(t) for t in transactions]
    state = SearchState(max_seconds, cancel_event=cancel_event,
                        progress_callback=progress_callback,
                        max_solutions=(max_results or MAX_SOLUTIONS))

    # Negative target or any negative transaction -> general path (two-sided pruning).
    # Otherwise -> the fast path based on the cents split (exploits all-positive values).
    if target_total <= 0 or any(c < 0 for c in tx_cents):
        return _solve_general(tx_cents, target_total, k_max, state)

    rounded, non_rounded, stats = preprocess(tx_cents, target_total)
    target_cents = target_total % 100
    n_nonround = len(non_rounded)
    n_round = len(rounded)
    # prefix of the round values (sorted descending): prefix_round[i] = sum of the i largest.
    prefix_round = [0] * (n_round + 1)
    for i in range(n_round):
        prefix_round[i + 1] = prefix_round[i] + rounded[i]

    total_candidates = 0
    try:
        # K = total number of items in the solution. c = how many are non-round; needed = round.
        for K in range(1, k_max + 1):
            for c in range(0, min(K, n_nonround) + 1):
                needed = K - c
                if needed > n_round:                # not enough round transactions to complete
                    continue
                # Tight window for the candidate's sum: the remainder is completed by exactly
                # `needed` round values, so rem is between (sum of the needed smallest) and
                # (sum of the needed largest).
                max_completion = prefix_round[needed]
                min_completion = prefix_round[n_round] - prefix_round[n_round - needed]
                lo = target_total - max_completion
                hi = target_total - min_completion
                cands = find_modular_candidates(
                    non_rounded, target_cents, c, state, lo, hi)
                for cand_values, cand_sum in cands:   # generator — consumed in O(K) memory
                    total_candidates += 1
                    rem = target_total - cand_sum
                    if rem < 0:
                        continue
                    if needed == 0:
                        if rem == 0:                # the candidate itself is a complete solution
                            state.record(cand_values, target_total)
                    else:
                        if rem == 0:
                            continue
                        assert rem % 100 == 0, "Remainder is not a whole unit — bug in the cents split!"
                        limit = state.max_solutions - len(state.solutions)
                        comps = find_subset_sum(rounded, rem, needed, state,
                                                exact_len=needed, limit=limit,
                                                prefix=prefix_round)
                        for comp in comps:
                            state.record(cand_values + comp, target_total)
                            if len(state.solutions) >= state.max_solutions:
                                break
                    if len(state.solutions) >= state.max_solutions:
                        break
                if len(state.solutions) >= state.max_solutions:
                    break
            if len(state.solutions) >= state.max_solutions:
                break
    except (_StopSearch, KeyboardInterrupt):
        if state.stop_reason is None:
            state.stop_reason = "interrupt"

    stats["n_candidates"] = total_candidates
    return _finalize(state, stats, target_total)


# --------------------------------------------------------------------------- #
#  Writing the Excel workbook (formatted input + formatted output).
# --------------------------------------------------------------------------- #
def write_input(filepath, transactions, target, k_max):
    """
    Creates a new, formatted Excel workbook with an 'Input' sheet ready for the algorithm:
    a "Transactions" header in column A (values from A2), "Target sum" in D1/D2,
    and the item limit in D3. Used for manual entry (no need to prepare an Excel file in advance).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.sheet_view.rightToLeft = False

    hdr_a = ws["A1"]
    hdr_a.value = "Transactions"
    hdr_a.font, hdr_a.fill, hdr_a.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER
    hdr_d = ws["D1"]
    hdr_d.value = "Target sum"
    hdr_d.font, hdr_d.fill, hdr_d.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER

    for i, v in enumerate(transactions, start=2):
        c = ws.cell(i, 1, float(v))
        c.alignment, c.number_format, c.border = _RIGHT, _MONEY, _BORDER

    tgt = ws["D2"]
    tgt.value = float(target)
    tgt.alignment, tgt.number_format, tgt.border = _RIGHT, _MONEY, _BORDER

    lbl = ws["C3"]
    lbl.value = "Max items per solution:"
    lbl.font, lbl.alignment = _MUTED_FONT, _RIGHT
    km = ws["D3"]
    km.value = int(k_max)
    km.alignment, km.border = _RIGHT, _BORDER

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16

    try:
        wb.save(filepath)
    except PermissionError:
        raise SystemExit("ERROR: the file is open/locked. Close it and try again.")


def write_output(filepath, solutions, stats):
    """Writes/replaces a formatted 'Output' sheet (RTL, headers, borders). Preserves the input sheet."""
    wb = openpyxl.load_workbook(filepath)          # without data_only — preserves formulas on save
    if "Output" in wb.sheetnames:
        del wb["Output"]
    ws = wb.create_sheet("Output")
    ws.sheet_view.rightToLeft = False

    ws["A1"] = "Reconciliation results"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")

    summary = [
        f'Total solutions found: {len(solutions)}',
        f"Status: {stats['status_text']}",
        (f"Transactions after filtering: {stats['kept']} "
         f"(round {stats['n_rounded']}, non-round {stats['n_non_rounded']}); "
         f"cents candidates: {stats['n_candidates']}; "
         f"branches examined: {stats['nodes']:,}; time: {stats['elapsed']:.1f}s"),
    ]
    for k, text in enumerate(summary):
        r = 2 + k
        ws.cell(r, 1, text).font = _LABEL_FONT if k == 0 else _MUTED_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    hdr_row = 6
    headers = ["Solution #", "# items", "Transactions", "Sum (check)"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(hdr_row, col, text)
        c.font, c.fill, c.alignment, c.border = _HEADER_FONT, _HEADER_FILL, _CENTER, _BORDER

    for idx, sol in enumerate(solutions, start=1):
        r = hdr_row + idx
        ws.cell(r, 1, idx)
        ws.cell(r, 2, len(sol))
        ws.cell(r, 3, format_amounts(sol))
        ws.cell(r, 4, sum(sol) / 100)
        for col in (1, 2, 3, 4):
            cell = ws.cell(r, col)
            cell.border, cell.alignment = _BORDER, _RIGHT
        ws.cell(r, 4).number_format = _MONEY

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 16

    try:
        wb.save(filepath)
    except PermissionError:
        raise SystemExit("ERROR: the file is open in Excel and locked for writing. Close it and run again.")


def append_analysis_sheet(filepath, action, value, remaining_solutions):
    """Adds a *new* sheet documenting one elimination step (ruling out / confirming a value)
    and the combinations that remain after it. Earlier sheets are preserved, so the workbook
    accumulates the entire elimination trail.

    action — "Eliminate" or "Confirm"; value — value in cents; remaining_solutions —
    list of tuples of cents that remain. Returns the name of the created sheet.
    """
    wb = openpyxl.load_workbook(filepath)
    n = 1                                          # free name: "Step 1", "Step 2", …
    while f"Step {n}" in wb.sheetnames:
        n += 1
    ws = wb.create_sheet(f"Step {n}")
    ws.sheet_view.rightToLeft = False

    verb = "ruled out" if action == "Eliminate" else "confirmed"
    sign = "-" if value < 0 else ""
    value_str = f"{sign}{abs(value) / 100:,.2f}"

    ws["A1"] = f"Analysis — Step {n}"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")

    summary = [
        f"Action: {verb} value {value_str}",
        f"Remaining combinations: {len(remaining_solutions)}",
    ]
    for k, text in enumerate(summary):
        r = 2 + k
        ws.cell(r, 1, text).font = _LABEL_FONT if k == 0 else _MUTED_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    hdr_row = 5
    headers = ["#", "# items", "Transactions", "Sum (check)"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(hdr_row, col, text)
        c.font, c.fill, c.alignment, c.border = _HEADER_FONT, _HEADER_FILL, _CENTER, _BORDER

    for idx, sol in enumerate(remaining_solutions, start=1):
        r = hdr_row + idx
        ws.cell(r, 1, idx)
        ws.cell(r, 2, len(sol))
        ws.cell(r, 3, format_amounts(sol))
        ws.cell(r, 4, sum(sol) / 100)
        for col in (1, 2, 3, 4):
            cell = ws.cell(r, col)
            cell.border, cell.alignment = _BORDER, _RIGHT
        ws.cell(r, 4).number_format = _MONEY

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 16

    try:
        wb.save(filepath)
    except PermissionError:
        raise SystemExit("ERROR: the file is open in Excel and locked for writing. Close it and try again.")
    return f"Step {n}"


# --------------------------------------------------------------------------- #
#  Run helpers.
# --------------------------------------------------------------------------- #
def _reconfig_stdout():
    """Ensures console output doesn't crash on the Windows console encoding."""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass


def _set_low_priority():
    """Lowers the process priority on Windows so the computer stays responsive. Returns True on success."""
    try:
        import ctypes
        BELOW_NORMAL_PRIORITY_CLASS = 0x00004000
        h = ctypes.windll.kernel32.GetCurrentProcess()
        return bool(ctypes.windll.kernel32.SetPriorityClass(h, BELOW_NORMAL_PRIORITY_CLASS))
    except Exception:
        return False


def _ask_minutes(default=5.0):
    """Asks the user for the time limit in minutes at the start of the run."""
    try:
        raw = input(f"Time limit in minutes (Enter = {default:g}; 0 = no limit): ").strip()
    except EOFError:
        raw = ""
    if raw == "":
        return default
    try:
        m = float(raw.replace(",", "."))
        return m if m >= 0 else default
    except ValueError:
        print(f"  Invalid input - using default {default:g} minutes.")
        return default


def main():
    _reconfig_stdout()
    parser = argparse.ArgumentParser(
        description="Accounting reconciliation tool (Subset Sum with branch-and-bound pruning)")
    parser.add_argument("filepath", help="path to the Excel file with an 'Input' sheet")
    parser.add_argument("--minutes", type=float, default=None,
                        help="time limit in minutes (0=no limit). If omitted — you'll be asked at runtime.")
    args = parser.parse_args()

    if sys.platform == "win32" and _set_low_priority():
        print("Process priority lowered (Below-Normal) - the computer stays responsive for other work.")

    print(f"Reading input from: {args.filepath}")
    transactions, target, k_max = read_input(args.filepath)
    print(f"Read {len(transactions)} transactions; target {target:.2f}; item limit {k_max}.")

    minutes = args.minutes if args.minutes is not None else _ask_minutes()
    if minutes and minutes > 0:
        max_seconds = minutes * 60
        print(f"Time limit: {minutes:g} min. You can stop anytime with Ctrl+C (found solutions are saved).")
    else:
        max_seconds = None
        print("Time limit: none. You can stop anytime with Ctrl+C (found solutions are saved).")

    print("Searching for solutions...")
    solutions, stats = solve(transactions, target, k_max, max_seconds)

    print()
    print(f"Found {len(solutions)} solutions. "
          f"{STATUS_TEXT_EN[stats['status']].format(n=stats['max_solutions'])}")
    for i, sol in enumerate(solutions, 1):
        print(f"  {i}. ({len(sol)} items) "
              + ", ".join(f"{c / 100:.2f}" for c in sol))

    write_output(args.filepath, solutions, stats)
    print(f"\nDone. Results written to the output sheet of: {args.filepath}")


if __name__ == "__main__":
    main()
