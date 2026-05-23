# -*- coding: utf-8 -*-
"""Creates a sample Excel file (example_reconcile.xlsx) for subset_sum_reconcile.py."""

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Input"

# Headers
ws["A1"] = "Transactions"
ws["D1"] = "Target sum"

# Target sum and item limit
ws["D2"] = 1000.50      # target with cents (.50)
ws["D3"] = 8            # maximum number of items in a solution

# Transaction list (column A from A2). Intentionally includes:
#   - solutions of varying lengths (1, 2, 3 items)
#   - duplicate values (two 400.00, two 50.00) to test duplicate prevention
#   - a transaction larger than the target (1500) that gets filtered out
#   - a negative value and text that get skipped
transactions = [
    1000.50,   # single-item solution
    600.50,    # 600.50 + 400.00  -> 2-item solution
    400.00,
    400.00,    # duplicate
    250.20,    # 250.20 + 350.30 + 400.00 -> 3-item solution
    350.30,
    777.77,    # noise
    333.00,    # noise
    12.34,     # noise
    999.99,    # noise (smaller than target, kept)
    1500.00,   # larger than target -> filtered out
    -5,        # negative -> skipped
    "abc",     # text -> skipped
    50.00,
    50.00,     # duplicate
]
for i, v in enumerate(transactions, start=2):
    ws.cell(i, 1, v)

wb.save("example_reconcile.xlsx")
print("Created example_reconcile.xlsx")
