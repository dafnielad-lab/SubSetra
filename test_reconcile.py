# -*- coding: utf-8 -*-
"""
Strategic regression test suite for the matching engine (subset_sum_reconcile.py).

Purpose: catch regressions after any code change. Just run:

    python test_reconcile.py

and read the PASS/FAIL summary. All tests must pass (0 FAIL).

The strongest test is the brute-force comparison: it generates hundreds of
random inputs and, for each, compares the engine's output against *every*
possible subset. If the engine misses a solution or invents one, the test
catches it immediately.
"""

import argparse
import datetime
import hashlib
import json
import os
import random
import tempfile
from itertools import combinations

import subset_sum_reconcile as core

_PASS = 0
_FAIL = 0


def check(name, ok, detail=""):
    global _PASS, _FAIL
    if ok:
        _PASS += 1
        print(f"  ok   {name}")
    else:
        _FAIL += 1
        print(f"  FAIL {name}   {detail}")


def engine_solutions(tx, target, k_max, max_results=10 ** 6):
    """Run the engine; return (set of cent-value tuples, original list)."""
    sols, _ = core.solve(tx, target, k_max, None, max_results=max_results)
    return set(tuple(s) for s in sols), sols


def brute(tx_cents, target_cents, k_max):
    """All subsets (up to k_max items) that sum exactly to target_cents, unique by values."""
    vals = [c for c in tx_cents if c != 0]
    out = set()
    for k in range(1, k_max + 1):
        for combo in combinations(range(len(vals)), k):
            if sum(vals[i] for i in combo) == target_cents:
                out.add(tuple(sorted((vals[i] for i in combo), reverse=True)))
    return out


# --------------------------------------------------------------------------- #
#  Generic comparison harness against brute-force.
# --------------------------------------------------------------------------- #
def run_bruteforce(label, gen_values, gen_target, n_cases, seed, max_brute=150):
    random.seed(seed)
    compared = 0
    for _ in range(n_cases):
        tx = gen_values()
        if len(tx) < 2:
            continue
        k_max = random.randint(2, 6)
        target = gen_target(tx, k_max)
        if target is None:
            continue
        tc = [core.to_cents(x) for x in tx]
        tgt = core.to_cents(target)
        b = brute(tc, tgt, k_max)
        if len(b) > max_brute:                 # skip cases with too many solutions
            continue
        e, sols = engine_solutions(tx, target, k_max)
        compared += 1

        if e != b:
            return check(label, False,
                         f"\n     tx={tx}\n     target={target} k_max={k_max}"
                         f"\n     brute={sorted(b)}\n     engine={sorted(e)}")
        for s in sols:                          # every solution sums to target, valid length
            if sum(s) != tgt or not (1 <= len(s) <= k_max):
                return check(label, False, f"invalid solution {s}")
        if len(sols) != len(e):                 # no duplicate solutions in the output
            return check(label, False, f"duplicate solutions in {sols}")

    check(f"{label}  ({compared} cases vs brute-force)", True)


# --------------------------------------------------------------------------- #
#  Brute-force test scenarios.
# --------------------------------------------------------------------------- #
def gen_positive():
    n = random.randint(5, 12)
    out = []
    for _ in range(n):
        v = round(random.uniform(0.01, 300), 2)
        if random.random() < 0.5:              # ~half whole-shekel (rounded pool vs agora)
            v = float(int(v)) or 1.0
        out.append(v)
    return out


def gen_mixed():
    n = random.randint(4, 10)
    return [x for x in (round(random.uniform(-200, 200), 2) for _ in range(n))
            if abs(x) >= 0.01]


def tgt_from_subset(tx, k_max):
    sub = random.sample(tx, random.randint(1, min(k_max, len(tx))))
    t = round(sum(sub), 2)
    return t if t > 0 else None


def tgt_from_subset_any(tx, k_max):
    sub = random.sample(tx, random.randint(1, min(k_max, len(tx))))
    t = round(sum(sub), 2)
    return t if abs(t) >= 0.01 else None


def tgt_zero(tx, k_max):
    return 0.0


# --------------------------------------------------------------------------- #
#  Targeted (strategic) scenarios.
# --------------------------------------------------------------------------- #
def test_agorot_precision():
    # 0.10 + 0.20 + 0.70 = 1.00  (classic floating-point inaccuracy trap)
    e, _ = engine_solutions([0.10, 0.20, 0.70, 0.55], 1.00, 4)
    check("agorot precision 0.10+0.20+0.70", (70, 20, 10) in e)
    # 33.33 * 3 = 99.99
    e, _ = engine_solutions([33.33, 33.33, 33.33, 0.01], 99.99, 5)
    check("agorot precision 33.33*3", (3333, 3333, 3333) in e)


def test_duplicates():
    # three 50s -> the combo {50,50} appears only once
    e, sols = engine_solutions([50.0, 50.0, 50.0], 100.0, 8)
    check("duplicates: {50,50} found once", (5000, 5000) in e and len(sols) == 1)
    # two rows of 50, target 50 -> one solution (not two)
    _, sols = engine_solutions([50.0, 50.0], 50.0, 8)
    check("duplicates: identical value-list shown once", len(sols) == 1)


def test_shortest_first():
    sols, _ = core.solve([100.0, 60.0, 40.0, 30.0, 70.0, 50.0, 1000.50],
                         100.0, 8, None, max_results=10)
    lengths = [len(s) for s in sols]
    check("shortest-first ordering", lengths == sorted(lengths), str(lengths))


def test_max_results_cap():
    tx = [float(v) for v in range(1, 30)]      # many combinations summing to 30
    sols, stats = core.solve(tx, 30.0, 8, None, max_results=3)
    check("max_results cap = 3", len(sols) == 3 and stats["status"] == "cap")
    # and still the shortest ones
    lengths = [len(s) for s in sols]
    check("cap returns shortest", lengths == sorted(lengths), str(lengths))


def test_powers_of_two():
    powers = [(2 ** k) / 100 for k in range(20)]   # 1..2^19 cents (all non-round)
    idx = [0, 2, 5, 7, 9, 11, 13, 16, 18]          # a unique subset
    target = round(sum(powers[i] for i in idx), 2)
    sols, stats = core.solve(powers, target, 12, 5)
    want = sorted((core.to_cents(powers[i]) for i in idx), reverse=True)
    check("2^N unique subset found", want in [sorted(s, reverse=True) for s in sols])
    check("2^N pruned hard (nodes < 1e6)", stats["nodes"] < 1_000_000,
          f"nodes={stats['nodes']}")


def test_super_increasing():
    # "Greedy" (super-increasing) sequences: each item exceeds the sum of all
    # smaller ones. Every target has a *unique* subset (like 2^N), which the
    # greedy largest-first method finds. We verify the engine returns exactly
    # that unique subset AND that the search finished (status=complete) -- i.e.
    # it pruned well and did not stall.
    random.seed(7)
    for _ in range(50):
        n = random.randint(6, 16)
        vals, total = [], 0
        for _ in range(n):
            nxt = total + random.randint(1, 25)         # > sum of all previous
            vals.append(nxt)
            total += nxt
        if random.random() < 0.5:
            vals = [v * 100 for v in vals]              # whole-shekel variant
        k_max = random.randint(2, 8)
        size = random.randint(1, min(k_max, n))
        idx = random.sample(range(n), size)
        tcents = sum(vals[i] for i in idx)
        sols, stats = core.solve([v / 100 for v in vals], tcents / 100, k_max, 30)
        want = sorted((vals[i] for i in idx), reverse=True)
        got = [sorted(s, reverse=True) for s in sols]
        if got != [want] or stats["status"] != "complete":
            return check("super-increasing (greedy) unique + complete", False,
                         f"n={n} k={k_max} want={want} got={got} status={stats['status']}")
    check("super-increasing (greedy) unique + complete (50 cases)", True)


def test_powers_of_two_varied():
    # 2^N varied: cents and whole-shekels, different subset sizes. Always unique.
    random.seed(11)
    cases = 0
    for whole in (False, True):
        for n in (10, 16, 22):
            base = [(2 ** k) * (100 if whole else 1) for k in range(n)]
            for size in (1, n // 2):
                k_max = max(size, 2)
                idx = random.sample(range(n), size)
                tcents = sum(base[i] for i in idx)
                sols, stats = core.solve([b / 100 for b in base], tcents / 100, k_max, 30)
                want = sorted((base[i] for i in idx), reverse=True)
                got = [sorted(s, reverse=True) for s in sols]
                cases += 1
                if got != [want] or stats["status"] != "complete":
                    return check("2^N varied", False,
                                 f"whole={whole} n={n} size={size} want={want} "
                                 f"got={got} status={stats['status']}")
    # edge case: the full set (target = sum of everything)
    base = [(2 ** k) for k in range(10)]
    sols, _ = core.solve([b / 100 for b in base], sum(base) / 100, 10, 30)
    if [sorted(s, reverse=True) for s in sols] != [sorted(base, reverse=True)]:
        return check("2^N varied (full set)", False, "full-set case failed")
    check(f"2^N varied (whole+agora, sizes, full-set; {cases + 1} cases)", True)


def test_target_zero_cancellation():
    sols, _ = core.solve([500.0, -500.0, 300.0, -300.0, 120.0], 0.0, 6, None)
    canon = [sorted(s, reverse=True) for s in sols]
    check("target=0 finds +500/-500", [50000, -50000] in canon)
    check("target=0 shortest first", sols and len(sols[0]) == 2)


def test_no_solution_short():
    # Short no-solution cases -- must be pruned immediately: 0 solutions, complete, fast.
    cases = [
        ([10.0, 20.0, 30.0], 15.50, "agora target from whole-shekel values"),
        ([0.02, 0.04, 0.06], 0.05, "odd-cent target from even-cent values"),
        ([1.0, 2.0, 3.0], 100.0, "target above total sum"),
        ([10.0, 20.0, 30.0], 5.0, "target below every item"),
        ([10.0, 20.0, 33.0], 7.0, "small unreachable"),
    ]
    for tx, target, label in cases:
        b = brute([core.to_cents(x) for x in tx], core.to_cents(target), 8)  # sanity
        sols, stats = core.solve(tx, target, 8, 5)
        ok = (len(b) == 0 and len(sols) == 0
              and stats["status"] == "complete" and stats["nodes"] < 100_000)
        if not ok:
            return check("short no-solution", False,
                         f"[{label}] sols={len(sols)} status={stats['status']} "
                         f"nodes={stats['nodes']} brute={len(b)}")
    check(f"short no-solution ({len(cases)} cases: empty, complete, fast)", True)


def test_no_solution_long():
    # A heavy search that provably has no solution: every value is a multiple of
    # 3 cents, and the target is congruent to 1 (mod 3) -- so no subset can sum
    # to it. The values are close in size, so pruning is weak and the search is
    # large; the time limit guarantees the test does not hang.
    random.seed(123)
    vals = sorted({3 * random.randint(8000, 14000) for _ in range(50)}, reverse=True)
    base = sum(random.sample(vals, 6))                  # a real sum (multiple of 3)
    target_cents = base + 1                             # +1 -> = 1 mod 3 -> unreachable
    tx = [v / 100 for v in vals]
    sols, stats = core.solve(tx, target_cents / 100, 12, 2.0)   # 2-second time limit
    check("long no-solution: 0 solutions (guaranteed)", len(sols) == 0)
    check("long no-solution: returns gracefully",
          stats["status"] in ("time", "complete"), f"status={stats['status']}")
    check("long no-solution: did NOT hang (time bounded)",
          stats["elapsed"] <= 8.0, f"elapsed={stats['elapsed']:.1f}s")


def test_excel_roundtrip():
    path = os.path.join(tempfile.gettempdir(), "_recon_test.xlsx")
    try:
        core.write_input(path, [1000.50, 600.50, 400.00, -5.0, 50.0], 1000.50, 7)
        tx, target, k_max = core.read_input(path)
        check("excel roundtrip: read target", abs(target - 1000.50) < 1e-9)
        check("excel roundtrip: read k_max", k_max == 7)
        check("excel roundtrip: read values", len(tx) == 5)
        sols, stats = core.solve(tx, target, k_max, None)
        core.write_output(path, sols, stats)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        # "Input" / "Output" are the actual sheet names the engine reads/writes.
        check("excel roundtrip: both sheets", set(wb.sheetnames) == {"Input", "Output"})
        check("excel roundtrip: input sheet LTR", not wb["Input"].sheet_view.rightToLeft)
        check("excel roundtrip: output sheet LTR", not wb["Output"].sheet_view.rightToLeft)
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_input_validation():
    import openpyxl

    def make(d2=None, d3=8, avals=(10,)):
        path = os.path.join(tempfile.gettempdir(), "_recon_bad.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Input"   # required sheet name
        for i, v in enumerate(avals, start=2):
            ws.cell(i, 1, v)
        if d2 is not None:
            ws["D2"] = d2
        ws["D3"] = d3
        wb.save(path)
        return path

    def expect_error(path, label):
        try:
            core.read_input(path)
            check(label, False, "no error raised")
        except SystemExit:
            check(label, True)
        finally:
            if os.path.exists(path):
                os.remove(path)

    expect_error(make(d2=None), "validation: empty target errors")
    expect_error(make(d2="abc"), "validation: non-numeric target errors")
    expect_error(make(d2=100, avals=()), "validation: no transactions errors")
    # target 0 is *allowed* (cancellations) -- must NOT raise
    p = make(d2=0, avals=(500, -500))
    try:
        core.read_input(p)
        check("validation: target=0 allowed", True)
    except SystemExit as e:
        check("validation: target=0 allowed", False, str(e))
    finally:
        if os.path.exists(p):
            os.remove(p)


def test_robust_under_timelimit():
    # Hard case (near-equal values) -- the test: no crash, returns gracefully
    # within a short time limit.
    random.seed(99)
    tx = [round(1000.50 + random.uniform(-5, 5), 2) for _ in range(40)]
    try:
        sols, stats = core.solve(tx, round(sum(tx[:14]), 2), 14, 0.5)   # 0.5 seconds
        check("robust: returns under time limit (no crash)",
              stats["status"] in ("complete", "cap", "time", "interrupt"))
    except Exception as e:                       # no unhandled exception allowed
        check("robust: returns under time limit (no crash)", False, repr(e))


# --------------------------------------------------------------------------- #
def run_deterministic():
    print("=" * 60)
    print(" Test suite -- matching engine")
    print("=" * 60)

    print("\n[1] Brute-force comparison (exhaustive search):")
    run_bruteforce("positive (mixed round/agora)", gen_positive,
                   tgt_from_subset, 1200, seed=1)
    run_bruteforce("mixed-sign (credits/debits)", gen_mixed,
                   tgt_from_subset_any, 1000, seed=2)
    run_bruteforce("target = 0 (cancellations)", gen_mixed,
                   tgt_zero, 1000, seed=3)

    print("\n[2] Targeted scenarios:")
    test_agorot_precision()
    test_duplicates()
    test_shortest_first()
    test_max_results_cap()
    test_powers_of_two()
    test_powers_of_two_varied()
    test_super_increasing()
    test_target_zero_cancellation()
    test_no_solution_short()
    test_no_solution_long()

    print("\n[3] Excel I/O and validation:")
    test_excel_roundtrip()
    test_input_validation()

    print("\n[4] Robustness:")
    test_robust_under_timelimit()

    print("\n[5] Saved failure corpus (past bugs kept as permanent regressions):")
    c_total, c_passed, c_fail = replay_corpus()
    if c_total == 0:
        print("  ok   corpus empty (no past failures recorded yet)")
    else:
        check(f"saved failure corpus ({c_total} past cases)", c_passed == c_total,
              f"{c_total - c_passed} regressed; first={c_fail}")

    print("\n" + "=" * 60)
    print(f"  Summary:  {_PASS} passed,  {_FAIL} failed")
    print("=" * 60)
    return 0 if _FAIL == 0 else 1


# --------------------------------------------------------------------------- #
#  Fuzz mode + cumulative log: accumulate (effectively-unique) random cases
#  over time. The log resets the "since last update" counters whenever the
#  engine source changes (detected by its hash).
# --------------------------------------------------------------------------- #
_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_log.json")
_ENGINE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "subset_sum_reconcile.py")
# Failure corpus: every case that ever failed is saved here and replayed as a
# permanent regression on every run -> bugs found once never silently return.
_CORPUS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "fuzz_failures.jsonl")


def _engine_hash():
    with open(_ENGINE_PATH, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()[:16]


def _load_log():
    try:
        with open(_LOG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _update_log(compared, passed):
    log = _load_log()
    now = datetime.datetime.now().isoformat(timespec="seconds")
    h = _engine_hash()
    log.setdefault("first_run", now)
    log.setdefault("total_cases", 0)
    log.setdefault("total_passed", 0)
    if log.get("code_hash") != h:                  # engine changed -> reset "since update"
        log["code_hash"] = h
        log["code_changed_at"] = now
        log["since_update_cases"] = 0
        log["since_update_passed"] = 0
    log["total_cases"] += compared
    log["total_passed"] += passed
    log["since_update_cases"] = log.get("since_update_cases", 0) + compared
    log["since_update_passed"] = log.get("since_update_passed", 0) + passed
    log["runs"] = log.get("runs", 0) + 1
    log["last_run"] = now
    with open(_LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(log, f, indent=2)
    return log


def _load_corpus():
    out = []
    if os.path.exists(_CORPUS_PATH):
        with open(_CORPUS_PATH, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        out.append(json.loads(line))
                    except Exception:
                        pass
    return out


def _fingerprint(tx, target, k_max):
    key = (tuple(sorted(core.to_cents(x) for x in tx)), core.to_cents(target), k_max)
    return hashlib.sha256(repr(key).encode()).hexdigest()[:16]


def _save_failures(failures):
    """Append new (deduped) failing cases to the corpus. Returns how many were added."""
    if not failures:
        return 0
    existing = {_fingerprint(e["tx"], e["target"], e["k_max"]) for e in _load_corpus()}
    added = 0
    with open(_CORPUS_PATH, "a", encoding="utf-8") as f:
        for fc in failures:
            fp = _fingerprint(fc["tx"], fc["target"], fc["k_max"])
            if fp in existing:
                continue
            existing.add(fp)
            f.write(json.dumps(fc, ensure_ascii=False) + "\n")
            added += 1
    return added


def replay_corpus():
    """Re-run every saved past failure; verify it now passes. Returns (total, passed, first_fail)."""
    total = passed = 0
    first_fail = None
    for entry in _load_corpus():
        tx, target, k_max = entry.get("tx"), entry.get("target"), entry.get("k_max")
        if tx is None or target is None or k_max is None:
            continue
        tgt = core.to_cents(target)
        tcents = [core.to_cents(x) for x in tx]
        nz = sum(1 for c in tcents if c != 0)
        if nz <= 14 and k_max <= 7:                  # exact: must match brute-force
            b = brute(tcents, tgt, k_max)
            e, sols = engine_solutions(tx, target, k_max)
            ok = (e == b and _valid(sols, tgt, k_max) and len(sols) == len(e))
        else:                                        # large: correctness only
            sols, _ = core.solve(tx, target, k_max, 1.0, max_results=10)
            ok = _valid(sols, tgt, k_max)
        total += 1
        if ok:
            passed += 1
        elif first_fail is None:
            first_fail = entry
    return total, passed, first_fail


# --- Fuzz case generators: a diverse zoo of shapes + occasional edge cases. ---
# Each returns (tx, target, k_max) or None to skip.
def _subset_target(tx, k_max, allow_zero=False, negative=False):
    if len(tx) < 1:
        return None
    for _ in range(10):
        t = round(sum(random.sample(tx, random.randint(1, min(k_max, len(tx))))), 2)
        if negative and t >= -0.01:
            continue
        if not allow_zero and abs(t) < 0.01:
            continue
        return t
    return None


def _valid(sols, tgt, k_max):
    """Correctness checks valid for ANY case: each sums to target, legal length, no dupes."""
    seen = set()
    for s in sols:
        if sum(s) != tgt or not (1 <= len(s) <= k_max):
            return False
        canon = tuple(sorted(s, reverse=True))
        if canon in seen:
            return False
        seen.add(canon)
    return True


def fc_positive_small():
    n = random.randint(4, 12)
    tx = [(float(int(v)) or 1.0) if random.random() < 0.5 else v
          for v in (round(random.uniform(0.01, 200), 2) for _ in range(n))]
    k = random.randint(2, 6)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fc_positive_large():
    tx = [round(random.uniform(100, 60000), 2) for _ in range(random.randint(4, 10))]
    k = random.randint(2, 6)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fc_mixed_sign():
    tx = [x for x in (round(random.uniform(-300, 300), 2) for _ in range(random.randint(4, 11)))
          if abs(x) >= 0.01]
    k = random.randint(2, 6)
    t = _subset_target(tx, k, allow_zero=True)
    return (tx, t, k) if (t is not None and len(tx) >= 2) else None


def fc_target_zero():
    tx = [x for x in (round(random.uniform(-300, 300), 2) for _ in range(random.randint(4, 11)))
          if abs(x) >= 0.01]
    return (tx, 0.0, random.randint(2, 6)) if len(tx) >= 2 else None


def fc_big_rounded_pool():
    tx = [float(random.randint(1, 5000)) for _ in range(random.randint(8, 18))]
    tx += [round(random.uniform(0.01, 50), 2) for _ in range(random.randint(2, 5))]
    k = random.randint(2, 6)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fc_near_equal():
    base = random.uniform(100, 2000)
    tx = [round(base + random.uniform(-3, 3), 2) for _ in range(random.randint(6, 12))]
    k = random.randint(2, 5)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fc_same_cents():
    cents = random.randint(1, 99) / 100
    tx = [round(random.randint(1, 500) + cents, 2) for _ in range(random.randint(5, 12))]
    k = random.randint(2, 6)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fc_tiny_agorot():
    tx = [round(random.uniform(0.01, 0.99), 2) for _ in range(random.randint(5, 12))]
    k = random.randint(2, 6)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fc_duplicates_heavy():
    distinct = [round(random.uniform(1, 100), 2) for _ in range(random.randint(2, 4))]
    tx = [random.choice(distinct) for _ in range(random.randint(6, 12))]
    k = random.randint(2, 6)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fc_super_increasing():
    cents, total = [], 0
    for _ in range(random.randint(6, 14)):
        nxt = total + random.randint(1, 20)
        cents.append(nxt)
        total += nxt
    if random.random() < 0.5:
        cents = [c * 100 for c in cents]
    k = random.randint(2, 8)
    tcents = sum(random.sample(cents, random.randint(1, min(k, len(cents)))))
    return ([c / 100 for c in cents], tcents / 100, k)


def fc_powers_two():
    n = random.randint(8, 16)
    base = [(2 ** i) * (100 if random.random() < 0.5 else 1) for i in range(n)]
    k = random.randint(2, 8)
    tcents = sum(random.sample(base, random.randint(1, min(k, n))))
    return ([b / 100 for b in base], tcents / 100, k)


def fe_single_item():
    tx = [round(random.uniform(0.01, 500), 2) for _ in range(random.randint(3, 10))]
    return (tx, random.choice(tx), random.randint(2, 6))


def fe_full_set():
    n = random.randint(2, 7)
    tx = [round(random.uniform(0.01, 100), 2) for _ in range(n)]
    return (tx, round(sum(tx), 2), n)


def fe_no_solution():
    cents = [3 * random.randint(1, 3000) for _ in range(random.randint(5, 12))]
    k = random.randint(2, 6)
    base = sum(random.sample(cents, min(k, len(cents))))
    return ([c / 100 for c in cents], (base + 1) / 100, k)   # +1 -> not a multiple of 3


def fe_min_items():
    tx = [round(random.uniform(0.01, 500), 2) for _ in range(random.randint(2, 3))]
    t = _subset_target(tx, len(tx))
    return (tx, t, len(tx)) if t else None


def fe_extreme():
    tx = [round(random.uniform(0.01, 0.5), 2) for _ in range(random.randint(4, 8))]
    tx.append(round(random.uniform(50000, 200000), 2))
    k = random.randint(2, 6)
    t = _subset_target(tx, k)
    return (tx, t, k) if t else None


def fe_negative_target():
    tx = [x for x in (round(random.uniform(-300, 300), 2) for _ in range(random.randint(5, 11)))
          if abs(x) >= 0.01]
    k = random.randint(2, 6)
    t = _subset_target(tx, k, negative=True)
    return (tx, t, k) if t else None


def fl_large():
    n = random.randint(22, 45)
    s = random.random()
    if s < 0.4:
        tx = [round(random.uniform(0.01, 5000), 2) for _ in range(n)]
    elif s < 0.7:
        tx = [float(random.randint(1, 9000)) for _ in range(n)]
    else:
        base = random.uniform(500, 5000)
        tx = [round(base + random.uniform(-10, 10), 2) for _ in range(n)]   # near-equal (hard)
    k = random.randint(4, 12)
    t = _subset_target(tx, k)            # a subset target -> a solution is planted (exists)
    return (tx, t, k) if t else None


# Weighted pool: common shapes more often, edge/large cases occasionally.
_FUZZ_GENS = (
    [fc_positive_small] * 10 + [fc_positive_large] * 5 + [fc_mixed_sign] * 8 +
    [fc_target_zero] * 6 + [fc_big_rounded_pool] * 6 + [fc_near_equal] * 5 +
    [fc_same_cents] * 4 + [fc_tiny_agorot] * 4 + [fc_duplicates_heavy] * 4 +
    [fc_super_increasing] * 4 + [fc_powers_two] * 4 +
    [fe_single_item] * 2 + [fe_full_set] * 2 + [fe_no_solution] * 3 +
    [fe_min_items] * 2 + [fe_extreme] * 2 + [fe_negative_target] * 2 +
    [fl_large] * 3
)


def run_fuzz(n_cases, seed=None):
    if seed is None:
        seed = random.randrange(2 ** 31)
    random.seed(seed)
    print("=" * 60)
    print(" FUZZ MODE -- diverse random cases + occasional edge/large cases")
    print("=" * 60)
    print(f"  seed = {seed}   (reproduce with:  --fuzz {n_cases} --seed {seed})")
    print(f"  generator shapes: {len(set(_FUZZ_GENS))}")

    # First, replay every past failure (permanent regressions).
    c_total, c_passed, _ = replay_corpus()
    if c_total:
        tag = "" if c_passed == c_total else "   !! REGRESSED"
        print(f"  regression corpus: {c_passed}/{c_total} past failures still pass{tag}")

    compared = passed = failed = 0
    first_fail = None
    new_failures = []
    for _ in range(n_cases):
        case = random.choice(_FUZZ_GENS)()
        if case is None or case[1] is None or len(case[0]) < 1:
            continue
        tx, target, k_max = case
        tgt = core.to_cents(target)
        tcents = [core.to_cents(x) for x in tx]
        nz = sum(1 for c in tcents if c != 0)

        if nz <= 14 and k_max <= 7:                  # small -> exact oracle (brute-force)
            b = brute(tcents, tgt, k_max)
            if len(b) > 200:                         # too many to compare exactly
                continue
            e, sols = engine_solutions(tx, target, k_max)
            ok = (e == b and _valid(sols, tgt, k_max) and len(sols) == len(e))
            detail = ("exact vs brute", sorted(b))
        else:                                        # large -> partial oracle (correctness)
            sols, stats = core.solve(tx, target, k_max, 1.0, max_results=10)
            ok = _valid(sols, tgt, k_max)
            if stats["status"] == "complete" and len(sols) == 0:
                ok = False                           # a solution was planted -> must be found
            detail = ("partial", stats["status"])

        compared += 1
        if ok:
            passed += 1
        else:
            failed += 1
            if first_fail is None:
                first_fail = (tx, target, k_max, detail,
                              [sorted(s, reverse=True) for s in sols])
            if len(new_failures) < 100:        # capture the case for the permanent corpus
                new_failures.append({
                    "tx": tx, "target": target, "k_max": k_max, "seed": seed,
                    "oracle": detail[0],
                    "found_at": datetime.datetime.now().isoformat(timespec="seconds"),
                })

    print(f"\n  this run:  {compared} cases compared,  {passed} passed,  {failed} failed")
    if first_fail:
        tx, target, k_max, detail, got = first_fail
        print("  !! FAILURE (reproduce with the seed above):")
        print(f"     tx={tx}")
        print(f"     target={target}  k_max={k_max}")
        print(f"     oracle={detail}")
        print(f"     engine={got}")

    added = _save_failures(new_failures)
    if added:
        print(f"  >> saved {added} NEW failing case(s) to "
              f"{os.path.basename(_CORPUS_PATH)} -- now permanent regressions")

    log = _update_log(compared, passed)
    print("\n  --- cumulative log (test_log.json) ---")
    print(f"  unique cases tested since {log['first_run'][:10]}:  "
          f"{log['total_cases']:,}  ({log['total_passed']:,} passed)")
    print(f"  since last engine update ({log.get('code_changed_at', '?')[:10]}):  "
          f"{log['since_update_cases']:,} cases,  {log['since_update_passed']:,} passed")
    print("=" * 60)
    return 1 if (failed or c_passed != c_total) else 0


def main():
    parser = argparse.ArgumentParser(description="Reconciliation engine test suite")
    parser.add_argument("--fuzz", nargs="?", const=2000, type=int, metavar="N",
                        help="run N fresh random cases (default 2000) and update the log")
    parser.add_argument("--seed", type=int, default=None,
                        help="fixed seed for --fuzz (to reproduce a previous run)")
    args = parser.parse_args()
    if args.fuzz is not None:
        return run_fuzz(args.fuzz, args.seed)
    return run_deterministic()


if __name__ == "__main__":
    raise SystemExit(main())
